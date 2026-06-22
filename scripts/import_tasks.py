"""
Importación histórica de TAREAS-PENDIENTES.xlsx a PostgreSQL.

Estructura del Excel:
  Col 1-9  : datos de la tarea (título, prioridad, solicitado, responsable, área, estado, fechas)
  Col 10+  : observaciones por día de daily/comité (cabecera = fecha del meeting)

Lógica:
  - Cada fila con título → 1 registro en `tasks`
  - Cada celda no vacía en columnas 10+ → 1 registro en `task_observations`
  - Se crean usuarios y áreas que no existan en BD
  - Se usa ON CONFLICT para usuarios/áreas para ser idempotente

Requirements: pip install openpyxl psycopg2-binary bcrypt
Ejecutar desde la raíz del proyecto: py scripts/import_tasks.py
"""

import openpyxl
import psycopg2
import bcrypt
import re
import sys
from datetime import datetime, date

sys.stdout.reconfigure(encoding='utf-8')

# ─────────────────────────────────────────────────────────────
# 0. Configuración
# ─────────────────────────────────────────────────────────────
EXCEL_PATH = 'excels/TAREAS-PENDIENTES.xlsx'
SYSTEM_USER_ID = 5          # SAMUEL PINEDA LOPEZ — fallback cuando el usuario no se resuelve
DAILY_COLS_START = 10       # primera columna con observaciones de dailys

# Contraseña por defecto para usuarios nuevos (deben cambiarla)
_raw_password = b'Elite2025!'
_hashed = bcrypt.hashpw(_raw_password, bcrypt.gensalt(12)).decode('utf-8')
DEFAULT_PASSWORD_HASH = _hashed.replace('$2b$', '$2y$', 1)  # Laravel usa $2y$

# ─────────────────────────────────────────────────────────────
# 1. Leer .env y conectar a PostgreSQL
# ─────────────────────────────────────────────────────────────
env_vars = {}
try:
    with open('.env', 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                k, v = line.split('=', 1)
                env_vars[k.strip()] = v.strip().strip("'\"")
except Exception as e:
    print(f'Error leyendo .env: {e}')
    sys.exit(1)

try:
    conn = psycopg2.connect(
        host=env_vars['DB_HOST'], port=env_vars['DB_PORT'],
        database=env_vars['DB_DATABASE'],
        user=env_vars['DB_USERNAME'], password=env_vars['DB_PASSWORD']
    )
    cur = conn.cursor()
    print(f"Conectado a {env_vars['DB_HOST']}/{env_vars['DB_DATABASE']}")
except Exception as e:
    print(f'Error conectando a BD: {e}')
    sys.exit(1)

# ─────────────────────────────────────────────────────────────
# 2. Verificar estado actual
# ─────────────────────────────────────────────────────────────
cur.execute('SELECT COUNT(*) FROM tasks')
existing_tasks = cur.fetchone()[0]
if existing_tasks > 0:
    print(f'\n[AVISO] Ya existen {existing_tasks} tarea(s) en la BD.')
    print('  El script AÑADIRÁ las del Excel (no borra las existentes).')
    resp = input('  ¿Continuar? (s/n): ').strip().lower()
    if resp != 's':
        print('Cancelado.')
        cur.close()
        conn.close()
        sys.exit(0)

# ─────────────────────────────────────────────────────────────
# 3. Helpers
# ─────────────────────────────────────────────────────────────
MONTHS_ES = {
    'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
    'mayo': 5, 'junio': 6, 'juniio': 6,  # typo en col 145
    'julio': 7, 'agosto': 8,
    'septiembre': 9, 'octubre': 10, 'noviembre': 11,
    'diciembre': 12, 'diciciembre': 12, 'dicicembre': 12,  # typos en col 84
}

def parse_column_date(header: str, col_idx: int):
    """
    Extrae una fecha de cabeceras como:
      'Daily 5 mayo 2025'          → 2025-05-05
      'Comite Jurídico 14 abril 2025' → 2025-04-14
      'Daily 6 de Noviembre'       → 2024-11-06  (sin año, infiere por posición)
    """
    if not header:
        return None
    text = re.sub(r'\s+', ' ', header.strip().lower())

    # Buscar mes
    month = None
    for m_name, m_num in MONTHS_ES.items():
        if m_name in text:
            month = m_num
            break
    if month is None:
        return None

    # Extraer números del texto
    numbers = re.findall(r'\d+', text)
    year, day = None, None
    for num_str in numbers:
        num = int(num_str)
        if num > 2000:
            year = num
        elif num <= 31 and day is None:
            day = num

    # Inferir año por posición de columna cuando no está explícito
    if year is None:
        if col_idx == 13:           # "Daily 6 de Noviembre" → antes de mayo 2025, retroactivo Nov 2024
            year = 2024
        elif col_idx <= 86:
            year = 2025
        else:
            year = 2026

    # Cols 87-88 dicen "Enero 2025" pero vienen después de Diciembre 2025
    if col_idx in (87, 88) and year == 2025 and month == 1:
        year = 2026

    if day is None:
        day = 1  # fallback al primer día del mes

    try:
        return date(year, month, day)
    except ValueError:
        return date(year, month, 1)


def norm(s: str) -> str:
    """Normaliza un nombre para comparación: mayúsculas, sin espacios dobles."""
    return re.sub(r'\s+', ' ', s.strip().upper()) if s else ''


def to_date(val):
    """Convierte datetime o date a date, devuelve None si es None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


# ─────────────────────────────────────────────────────────────
# 4. Preparar usuarios
# ─────────────────────────────────────────────────────────────
print('\n=== PREPARANDO USUARIOS ===')

# Mapeo directo: nombre normalizado del Excel → user_id en BD
# Cubre variaciones de escritura, apodos, etc.
STATIC_USER_MAP = {
    'CARTERA INVERCONSTRUCCION':    7,   # LUZ NELLY MORALES VALENCIA
    'NELLY MORALES':                7,
    'EDGAR ANDRES LANCE':           23,
    'INGRID OSPINO':                2,   # INGRID PAOLA OSPICIO PACHECO (typo en Excel)
    'INGRID OSPICIO':               2,
    'LAURA G':                      6,   # LAURA GOMEZ
    'LAURA GOMEZ':                  6,
    'LAURA GÓMEZ':                  6,
    'LAURA GÓMEZ CORREA':           6,
    'MANUELA MEJIA':                32,  # MANUELA MARIA MEJIA GOMEZ
    'SANTIAGO SANCHEZ':             9,
    'SANTIAGO SÁNCHEZ':             9,
    'SSANCHEZ@CYBPROJECT.COM':      9,   # mismo usuario, email externo
    'SARA':                         1,   # SARA ELENA MORENO OROZCO
    'SARA MORENO':                  1,
    'JULIÁN POSADA':                25,  # JULIAN ANDRÉS POSADA MORALES
    'JULIAN POSADA':                25,
    'MARIA CAROLINA HOYOS SEJIN':   31,
    'MARIA CAROLINA HOYOS SEJÍN':   31,
}

# Usuarios nuevos a crear: (clave_normalizada, nombre_completo, email)
# Las claves son todas las variaciones que aparecen en el Excel.
NEW_USERS_SPEC = [
    # variaciones de Juliana Arango → un solo usuario
    ('JULIANA ARANGO MARIN',         'Juliana Arango Marin',                  'juliana.arango@inverconstruccion.com'),
    ('JULIANA ARANGO',               'Juliana Arango Marin',                  'juliana.arango@inverconstruccion.com'),
    ('JULIAN ARANGO',                'Juliana Arango Marin',                  'juliana.arango@inverconstruccion.com'),
    # otros
    ('LADY GAONA',                   'Lady Gaona',                            'lady.gaona@elite.com'),
    ('FELIPE ACOSTA',                'Felipe Acosta',                         'felipe.acosta@elite.com'),
    ('FELIPE ARIAS',                 'Felipe Arias',                          'felipe.arias@elite.com'),
    ('GERENCIA',                     'Gerencia Elite',                        'gerencia@elite.com'),
    ('NICOLAS QUICENO',              'Nicolas Quiceno',                       'nicolas.quiceno@elite.com'),
    ('RRHH',                         'Recursos Humanos',                      'rrhh@elite.com'),
    ('RECURSOS HUMANOS',             'Recursos Humanos',                      'rrhh@elite.com'),
    ('SANTIAGO BARON',               'Santiago Baron',                        'santiago.baron@elite.com'),
    ('SOFIA YEPES',                  'Sofia Yepes',                           'sofia.yepes@elite.com'),
    ('DIRECCIONCOMERCIAL@INVERCONSTRUCCION.COM', 'Dirección Comercial Inverconstrucción', 'direccioncomercial@inverconstruccion.com'),
    ('GERENTEDEPROYECTOS@INVERCONSTRUCCION.COM', 'Gerente de Proyectos Inverconstrucción','gerentedeproyectos@inverconstruccion.com'),
    ('AREA JURIDICA INVERCONSTRUCCIÓN S.A.S.',  'Área Jurídica Inverconstrucción',       'juridica@inverconstruccion.com'),
    ('AREA JURIDICA INVERCONSTRUCCION S.A.S.',  'Área Jurídica Inverconstrucción',       'juridica@inverconstruccion.com'),
    ('RESIDENTE ESTRUCTURA',         'Residente Estructura',                  'residente.estructura@elite.com'),
    ('TRÁMITES CIUDADELA SAN MIGUEL','Trámites Ciudadela San Miguel',         'tramites@inverconstruccion.com'),
    ('TRAMITES CIUDADELA SAN MIGUEL','Trámites Ciudadela San Miguel',         'tramites@inverconstruccion.com'),
    ('VALERIA GIRALDO HENAO',        'Valeria Giraldo Henao',                 'valeria.giraldo@elite.com'),
    ('INVERCONSTRUCCIONSAS@GMAIL.COM','Inverconstrucción SAS',                'inverconstruccionsas@gmail.com'),
    ('MMEJIA@CYBPROJECT.COM',        'M. Mejia Cybproject',                   'mmejia@cybproject.com'),
]

# Cache: clave_normalizada → user_id
user_cache: dict[str, int] = dict(STATIC_USER_MAP)

# Cargar usuarios existentes en cache (por nombre y por email)
cur.execute('SELECT id, name, email FROM users')
for uid, uname, uemail in cur.fetchall():
    if uname:
        user_cache[norm(uname)] = uid
    if uemail:
        user_cache[norm(uemail)] = uid

# Crear los usuarios que falten (idempotente por email)
email_to_id: dict[str, int] = {}
for cache_key, full_name, email in NEW_USERS_SPEC:
    email_key = norm(email)

    # ¿Ya está en cache?
    if cache_key in user_cache:
        email_to_id[email] = user_cache[cache_key]
        continue

    # ¿El email ya existe en BD?
    if email_key in user_cache:
        user_cache[cache_key] = user_cache[email_key]
        email_to_id[email] = user_cache[email_key]
        continue

    # ¿Ya lo creamos en esta ejecución?
    if email in email_to_id:
        user_cache[cache_key] = email_to_id[email]
        continue

    # Insertar
    cur.execute(
        'INSERT INTO users (name, email, password, created_at, updated_at) '
        'VALUES (%s, %s, %s, NOW(), NOW()) ON CONFLICT (email) DO NOTHING RETURNING id',
        (full_name, email, DEFAULT_PASSWORD_HASH)
    )
    row = cur.fetchone()
    if row:
        new_id = row[0]
        print(f'  [NUEVO] {full_name} ({email}) → id={new_id}')
    else:
        # ON CONFLICT: recuperar el id existente
        cur.execute('SELECT id FROM users WHERE email = %s', (email,))
        new_id = cur.fetchone()[0]

    user_cache[cache_key] = new_id
    user_cache[email_key] = new_id
    email_to_id[email] = new_id


def resolve_user(raw) -> int | None:
    """Devuelve user_id a partir de un valor crudo del Excel (nombre o email)."""
    if not raw:
        return None
    s = str(raw).strip()
    key = norm(s)
    if key in user_cache:
        return user_cache[key]
    # Búsqueda parcial (por si hay typos menores)
    for ck, uid in user_cache.items():
        if ck and (key in ck or ck in key) and len(ck) > 4:
            return uid
    print(f'  [SIN MAPA] usuario: {repr(s)}')
    return None


# ─────────────────────────────────────────────────────────────
# 5. Preparar áreas
# ─────────────────────────────────────────────────────────────
print('\n=== PREPARANDO ÁREAS ===')

STATIC_AREA_MAP = {
    'CARTERA':         7,   # Trámites y Cartera
    'COMERCIAL':       1,
    'JURÍDICA':        8,
    'JURIDICA':        8,
    'TÉCNICA':         5,
    'TECNICA':         5,
}

NEW_AREAS_SPEC = [
    ('ESCRITURACIÓN',   'Escrituración'),
    ('ESCRITURACION',   'Escrituración'),
    ('GERENCIA',        'Gerencia'),
    ('GESTIÓN HUMANA',  'Gestión Humana'),
    ('GESTION HUMANA',  'Gestión Humana'),
    ('MARKETING',       'Marketing'),
]

area_cache: dict[str, int] = dict(STATIC_AREA_MAP)

# Cargar áreas existentes
cur.execute('SELECT id, name FROM areas')
for aid, aname in cur.fetchall():
    area_cache[norm(aname)] = aid

# Crear las que falten
area_name_to_id: dict[str, int] = {}
for cache_key, full_name in NEW_AREAS_SPEC:
    if cache_key in area_cache:
        area_name_to_id[full_name] = area_cache[cache_key]
        continue
    if full_name in area_name_to_id:
        area_cache[cache_key] = area_name_to_id[full_name]
        continue
    cur.execute(
        'INSERT INTO areas (name, created_at, updated_at) VALUES (%s, NOW(), NOW()) RETURNING id',
        (full_name,)
    )
    new_id = cur.fetchone()[0]
    area_cache[cache_key] = new_id
    area_cache[norm(full_name)] = new_id
    area_name_to_id[full_name] = new_id
    print(f'  [NUEVA] {full_name} → id={new_id}')


def resolve_area(raw) -> int | None:
    if not raw:
        return None
    key = norm(str(raw).strip())
    val = area_cache.get(key)
    if val is None:
        print(f'  [SIN MAPA] área: {repr(raw)}')
    return val


# ─────────────────────────────────────────────────────────────
# 6. Cargar Excel y parsear fechas de columnas
# ─────────────────────────────────────────────────────────────
print('\n=== CARGANDO EXCEL ===')
try:
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet = wb.active
    print(f'Hoja: {sheet.title} | Filas: {sheet.max_row} | Columnas: {sheet.max_column}')
except Exception as e:
    print(f'Error cargando Excel: {e}')
    cur.close()
    conn.close()
    sys.exit(1)

# Parsear fecha de cada columna de daily
column_dates: dict[int, date] = {}
for col_idx in range(DAILY_COLS_START, sheet.max_column + 1):
    header = sheet.cell(row=1, column=col_idx).value
    if header:
        parsed = parse_column_date(str(header), col_idx)
        column_dates[col_idx] = parsed

# Mostrar las que no se pudieron parsear
unparsed = [
    (col_idx, sheet.cell(row=1, column=col_idx).value)
    for col_idx, d in column_dates.items()
    if d is None
]
if unparsed:
    print(f'Columnas sin fecha parseada ({len(unparsed)}):')
    for ci, h in unparsed:
        print(f'  col {ci}: {repr(h)}')

# ─────────────────────────────────────────────────────────────
# 7. Importar tareas y observaciones
# ─────────────────────────────────────────────────────────────
print('\n=== IMPORTANDO ===')

STATUS_VALID = {'Por hacer', 'En espera', 'En progreso', 'Completada'}
PRIORITY_VALID = {'P0', 'P1', 'P2', 'P3'}

tasks_created = 0
obs_created = 0
rows_skipped = 0

try:
    for row_idx in range(2, sheet.max_row + 1):
        title_raw = sheet.cell(row=row_idx, column=1).value
        if not title_raw or not str(title_raw).strip():
            rows_skipped += 1
            continue

        title = str(title_raw).strip()[:255]  # max 255 chars (VARCHAR limit)

        priority_raw = sheet.cell(row=row_idx, column=2).value
        priority = str(priority_raw).strip() if priority_raw else 'P2'
        if priority not in PRIORITY_VALID:
            priority = 'P2'

        status_raw = sheet.cell(row=row_idx, column=6).value
        status = str(status_raw).strip() if status_raw else 'Por hacer'
        if status not in STATUS_VALID:
            status = 'Por hacer'

        requested_by_id = resolve_user(sheet.cell(row=row_idx, column=3).value) or SYSTEM_USER_ID
        responsible_id  = resolve_user(sheet.cell(row=row_idx, column=4).value)
        area_id         = resolve_area(sheet.cell(row=row_idx, column=5).value)

        start_date      = to_date(sheet.cell(row=row_idx, column=7).value)
        scheduled_end   = to_date(sheet.cell(row=row_idx, column=8).value)
        actual_end      = to_date(sheet.cell(row=row_idx, column=9).value)

        # Si está Completada pero sin fecha real, usamos la programada como mejor aproximación
        if status == 'Completada' and actual_end is None:
            actual_end = scheduled_end

        # Insertar tarea
        cur.execute(
            '''INSERT INTO tasks
               (title, priority, status, requested_by_id, responsible_id, area_id,
                start_date, scheduled_end_date, actual_end_date, created_at, updated_at)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
               RETURNING id''',
            (title, priority, status, requested_by_id, responsible_id, area_id,
             start_date, scheduled_end, actual_end)
        )
        task_id = cur.fetchone()[0]
        tasks_created += 1

        # El observador de las notas es el responsable, o el solicitante, o el sistema
        observer_id = responsible_id or requested_by_id or SYSTEM_USER_ID

        # Insertar observaciones de dailys
        for col_idx in range(DAILY_COLS_START, sheet.max_column + 1):
            obs_raw = sheet.cell(row=row_idx, column=col_idx).value
            if not obs_raw or not str(obs_raw).strip():
                continue

            obs_text = str(obs_raw).strip()
            obs_date = column_dates.get(col_idx)

            if obs_date:
                obs_ts = datetime(obs_date.year, obs_date.month, obs_date.day, 12, 0, 0)
            else:
                obs_ts = None

            cur.execute(
                '''INSERT INTO task_observations
                   (task_id, user_id, observation, created_at, updated_at)
                   VALUES (%s, %s, %s, %s, %s)''',
                (task_id, observer_id, obs_text,
                 obs_ts or datetime.now(), obs_ts or datetime.now())
            )
            obs_created += 1

        if tasks_created % 50 == 0:
            print(f'  ... {tasks_created} tareas insertadas')

    conn.commit()
    print(f'\n{"="*50}')
    print('IMPORTACIÓN COMPLETADA CON ÉXITO')
    print(f'  Tareas importadas:       {tasks_created}')
    print(f'  Observaciones importadas: {obs_created}')
    print(f'  Filas vacías omitidas:   {rows_skipped}')
    print(f'{"="*50}')
    print()
    print('IMPORTANTE: Los usuarios nuevos tienen contraseña "Elite2025!"')
    print('  Pídeles que la cambien desde el panel de administración.')

except Exception as e:
    conn.rollback()
    print(f'\n[ERROR] Se revirtieron todos los cambios: {e}')
    import traceback
    traceback.print_exc()
finally:
    cur.close()
    conn.close()

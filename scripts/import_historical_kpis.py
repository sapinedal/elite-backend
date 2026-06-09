import openpyxl
import psycopg2
import sys
import os

# Configurar codificación para salida en Windows
sys.stdout.reconfigure(encoding='utf-8')

# 1. Leer variables de entorno desde .env
env_vars = {}
try:
    with open(".env", "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                key, val = line.split("=", 1)
                # Remover comillas si existen
                val = val.strip("'\"")
                env_vars[key.strip()] = val
except Exception as e:
    print(f"Error al leer el archivo .env: {e}")
    sys.exit(1)

db_host = env_vars.get("DB_HOST", "127.0.0.1")
db_port = env_vars.get("DB_PORT", "5432")
db_name = env_vars.get("DB_DATABASE", "elite_backend")
db_user = env_vars.get("DB_USERNAME", "postgres")
db_pass = env_vars.get("DB_PASSWORD", "root")

print("Configuración de BD encontrada:")
print(f"Host: {db_host}, Puerto: {db_port}, BD: {db_name}, Usuario: {db_user}")

# 2. Conectar a PostgreSQL
try:
    conn = psycopg2.connect(
        host=db_host,
        port=db_port,
        database=db_name,
        user=db_user,
        password=db_pass
    )
    cursor = conn.cursor()
    print("Conexión a PostgreSQL establecida con éxito.")
except Exception as e:
    print(f"Error al conectar a PostgreSQL: {e}")
    sys.exit(1)

# 3. Mapear Cargos en Excel a IDs de Usuarios en BD
# DIRECTORA COMERCIAL -> User ID 1
# LIDER SALA DE VENTAS -> User ID 2
# ASESORA COMERCIAL NATALIA -> User ID 4
# ASESORA COMERCIAL PAOLA -> User ID 3
user_mapping = {
    "DIRECTORA COMERCIAL": 1,
    "LIDER SALA DE VENTAS": 2,
    "ASESORA COMERCIAL NATALIA": 4,
    "ASESORA COMERCIAL PAOLA": 3
}

# 4. Mapear columnas a Mes/Año
# AQ (43) -> Agosto 2025
# AR (44) -> Septiembre 2025
# AS (45) -> Octubre 2025
# AT (46) -> Noviembre 2025
# AU (47) -> Diciembre-Enero 2025-2026 (se insertan dos registros: Dic y Ene)
# AV (48) -> Febrero 2026
# AW (49) -> Marzo 2026
column_mappings = {
    43: [("Agosto", 8, 2025)],
    44: [("Septiembre", 9, 2025)],
    45: [("Octubre", 10, 2025)],
    46: [("Noviembre", 11, 2025)],
    47: [("Diciembre", 12, 2025), ("Enero", 1, 2026)],
    48: [("Febrero", 2, 2026)],
    49: [("Marzo", 3, 2026)]
}

# 5. Cargar el libro de Excel
try:
    excel_path = "KPI´s_Área comercial.xlsx"
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if "Tablero Mar-26" not in wb.sheetnames:
        print("La hoja 'Tablero Mar-26' no se encuentra en el archivo Excel.")
        sys.exit(1)
    sheet = wb["Tablero Mar-26"]
    print("Excel cargado correctamente.")
except Exception as e:
    print(f"Error al cargar el Excel: {e}")
    sys.exit(1)

# 6. Procesar filas y columnas
imported_count = 0
updated_count = 0

try:
    for row in range(2, 6): # Filas 2 a 5
        cargo_raw = sheet.cell(row=row, column=42).value # Columna AP (42)
        if not cargo_raw:
            continue
        cargo = cargo_raw.strip()
        user_id = user_mapping.get(cargo)
        
        if not user_id:
            print(f"Cargo no reconocido en fila {row}: '{cargo}'")
            continue
            
        print(f"\nProcesando cargo '{cargo}' (User ID: {user_id})...")
        
        for col_idx, periods in column_mappings.items():
            score_raw = sheet.cell(row=row, column=col_idx).value
            if score_raw is None or score_raw == '':
                # Si está vacío, no se registra nada
                continue
                
            try:
                score = float(score_raw)
            except ValueError:
                print(f"Puntaje inválido en fila {row}, col {col_idx}: {score_raw}")
                continue
                
            # Insertar o actualizar para cada mes asignado a esta columna
            for month_name, month, year in periods:
                print(f" - {month_name} ({month}/{year}): {score}")
                
                # Ejecutar INSERT ON CONFLICT DO UPDATE
                query = """
                    INSERT INTO evaluations (
                        user_id, month, year, total_score, status, general_analysis, created_at, updated_at
                    ) VALUES (%s, %s, %s, %s, %s, %s, NOW(), NOW())
                    ON CONFLICT (user_id, month, year)
                    DO UPDATE SET 
                        total_score = EXCLUDED.total_score,
                        status = EXCLUDED.status,
                        updated_at = NOW()
                    RETURNING (xmax = 0);
                """
                cursor.execute(query, (
                    user_id, 
                    month, 
                    year, 
                    score, 
                    'finalizada', 
                    'Importado históricamente desde Excel'
                ))
                
                is_insert = cursor.fetchone()[0]
                if is_insert:
                    imported_count += 1
                else:
                    updated_count += 1
                    
    # Confirmar transacción
    conn.commit()
    print("\n=========================================")
    print("PROCESO FINALIZADO CON ÉXITO")
    print(f"Evaluaciones creadas: {imported_count}")
    print(f"Evaluaciones actualizadas (sobrescritas): {updated_count}")
    print("=========================================")
    
except Exception as e:
    conn.rollback()
    print(f"\nError durante la importación (se revirtieron los cambios): {e}")
finally:
    cursor.close()
    conn.close()
